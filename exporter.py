"""
exporter.py — Book1-format Mouser BOM Excel exporter.

  Row 1 : Project title banner
  Row 2 : Disclaimer note
  Row 3 : Column headers  (frozen below)
  Row 4+: One row per unique component (grouped by MPN, or value+footprint
          for passives/components without an MPN)
  Last  : TOTAL cost row + colour legend

Columns (14):
  #  Reference  Qty  Description  Footprint  MPN  Manufacturer
  Vendor P/N (Mouser)
  Unit Price (₹)  Unit Price (₹ +18% GST)
  Ext. Price (₹)  Ext. Price (₹ +18% GST)
  Lead Time  Purchase Link

Row colour coding:
  Green  — Resistors, Capacitors, generic LEDs   (buy locally)
  Yellow — TVS/Schottky diodes, Fuses             (try local first)
  Blue   — ICs, MOSFETs, Connectors               (order from Mouser)
"""

from __future__ import annotations

import os
import re
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import GST_RATE


# ── Palette ───────────────────────────────────────────────────────────────────
_NAVY     = "1A2744"
_MID_NAVY = "2C4770"
_NOTE_BG  = "E8EDF5"
_LINK_CLR = "1155CC"

# Row zebra — base light/dark per category
_CAT_COLORS = {
    "green":  ("D9EAD3", "EAF4E3"),   # resistors, caps, LEDs
    "yellow": ("FFF2CC", "FFFDE7"),   # TVS, schottky, fuses
    "blue":   ("CFE2F3", "E3F2FD"),   # ICs, MOSFETs, connectors
    "gray":   ("F3F3F3", "FFFFFF"),   # everything else
}

# Legend text matching the colours
_LEGEND = [
    ("green",  "Green  = Resistors / Caps / generic LEDs  → buy locally"),
    ("yellow", "Yellow = TVS / Schottky diodes, Fuses      → try local, else Mouser"),
    ("blue",   "Blue   = ICs / MOSFET / Connectors         → order from Mouser"),
]


# ── Style helpers ─────────────────────────────────────────────────────────────

def _side(color: str = "C5CBD8") -> Side:
    return Side(style="thin", color=color)

def _border() -> Border:
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", start_color=hex_color)

def _font(size: int = 9, bold: bool = False, color: str = "000000",
          italic: bool = False) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Column definitions ────────────────────────────────────────────────────────
_COLS = [
    ("#",                                5,  "center"),
    ("Reference",                       22,  "left"),
    ("Qty",                              6,  "center"),
    ("Description",                     46,  "left"),
    ("Footprint",                       26,  "left"),
    ("MPN",                             26,  "left"),
    ("Manufacturer",                    22,  "left"),
    ("Vendor P/N (Mouser)",             26,  "left"),
    ("Unit Price\n(₹)",                 14,  "right"),   # col I
    ("Unit Price\n(₹ +18% GST)",        16,  "right"),   # col J
    ("Ext. Price\n(₹)",                 14,  "right"),   # col K
    ("Ext. Price\n(₹ +18% GST)",        16,  "right"),   # col L
    ("Lead Time",                       14,  "center"),
    ("Purchase Link",                   52,  "left"),
]

_UNIT_COL = "I"   # raw unit price — referenced in GST/ext formulas
_QTY_COL  = "C"
_EXT_COL  = "K"   # Ext. Price (₹) — summed in TOTAL row


# ── Component category detector ───────────────────────────────────────────────

def _category(ref: str, description: str, value: str) -> str:
    """Return 'green', 'yellow', 'blue', or 'gray' for row colouring."""
    r   = ref.strip().upper()
    d   = description.lower()
    v   = value.lower()
    pfx = re.match(r'^([A-Z]+)', r)
    p   = pfx.group(1) if pfx else ""

    if p in ("R",):
        return "green"
    if p == "C":
        return "green"
    if p == "D":
        if any(w in d or w in v for w in ("led", "light emitting")):
            return "green"
        if any(w in d or w in v for w in ("tvs", "transient", "schottky", "rectifier", "smbj", "smdj", "ss5", "ss54")):
            return "yellow"
        return "yellow"
    if p == "F":
        return "yellow"
    if p in ("U", "Q"):
        return "blue"
    if p in ("J", "P", "CN"):
        return "blue"
    return "gray"


# ── Grouping key ──────────────────────────────────────────────────────────────

_SYNTH_SEP = "\x00"   # invisible separator — never in real MPNs

def _group_key(row: pd.Series) -> str:
    """
    Return a stable grouping key for this BOM row.

    Priority:
      1. Real MPN from enrichment or BOM
      2. value + footprint  (keeps different resistor values separate)
      3. description fallback
    """
    mpn = str(row.get("mpn", "") or "").strip()
    if mpn and mpn.lower() not in ("nan", "none"):
        return mpn

    value = str(row.get("value", "") or "").strip()
    fp    = str(row.get("footprint", "") or "").strip()
    if value:
        return f"{value}{_SYNTH_SEP}{fp}"

    desc = str(row.get("description_enriched", "")
               or row.get("description", "")
               or "UNKNOWN").strip()
    return f"{desc}{_SYNTH_SEP}{fp}"


def _is_synthetic(key: str) -> bool:
    return _SYNTH_SEP in key


# ── Misc helpers ──────────────────────────────────────────────────────────────

def _safe_float(val) -> float | None:
    if val is None:
        return None
    try:
        cleaned = re.sub(r"[^\d.\-]", "", str(val).replace(",", "").strip())
        return float(cleaned) if cleaned else None
    except Exception:
        return None


def _lead_time(stock: int, qty: int) -> str:
    try:
        stock = int(stock or 0)
    except (ValueError, TypeError):
        stock = 0
    if stock == 0:
        return "Check Mouser"
    if stock >= qty * 3:
        return "In stock"
    return "1–2 weeks"


def _purchase_url(product_url: str, mouser_pn: str) -> str:
    if product_url and product_url.startswith("http"):
        return product_url
    pn = (mouser_pn or "").strip().replace(" ", "")
    if not pn or pn in ("nan", "None"):
        return ""
    return f"https://www.mouser.in/ProductDetail/{pn}"


def _fmt_footprint(fp: str) -> str:
    """Return just the package part of a KiCad footprint string."""
    if ":" in fp:
        fp = fp.split(":")[-1]
    return fp.strip()


# ── Sheet writer ──────────────────────────────────────────────────────────────

def _write_sheet(ws, df: pd.DataFrame, project_name: str) -> None:
    ws.sheet_view.showGridLines = False
    ncols    = len(_COLS)
    last_col = get_column_letter(ncols)
    gst_mult = 1 + GST_RATE

    # ── Row 1: title banner ───────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    c           = ws["A1"]
    c.value     = f"{project_name} — BOM (Mouser)"
    c.font      = _font(size=13, bold=True, color="FFFFFF")
    c.fill      = _fill(_NAVY)
    c.alignment = _align("left", "center")
    ws.row_dimensions[1].height = 24

    # ── Row 2: disclaimer ─────────────────────────────────────────────────────
    today = datetime.now().strftime("%B %Y")
    ws.merge_cells(f"A2:{last_col}2")
    c           = ws["A2"]
    c.value     = (
        f"Indicative prices (₹ INR, single-unit) — Mouser IN only. "
        f"GST column = base price × {gst_mult:.2f} (18% GST). "
        f"Verify stock & pricing before ordering. {today}."
    )
    c.font      = _font(size=8, italic=True, color="444444")
    c.fill      = _fill(_NOTE_BG)
    c.alignment = _align("left", "center")
    ws.row_dimensions[2].height = 16

    # ── Row 3: column headers ─────────────────────────────────────────────────
    white_b = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="thin", color="FFFFFF"),
    )
    for ci, (label, width, _) in enumerate(_COLS, 1):
        c           = ws.cell(row=3, column=ci, value=label)
        c.font      = _font(size=9, bold=True, color="FFFFFF")
        c.fill      = _fill(_MID_NAVY)
        c.alignment = _align("center", "center", wrap=True)
        c.border    = white_b
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[3].height = 30

    # ── Group BOM rows ────────────────────────────────────────────────────────
    seen:  dict[str, dict] = {}
    order: list[str]       = []

    for _, row in df.iterrows():
        key = _group_key(row)
        qty = int(row.get("qty", 1) or 1)
        ref = str(row.get("reference", ""))

        if key in seen:
            seen[key]["qty"]  += qty
            seen[key]["refs"]  = f"{seen[key]['refs']}, {ref}"
        else:
            fp_raw = str(row.get("footprint", "") or "")
            seen[key] = {
                "refs":         ref,
                "qty":          qty,
                "description":  str(row.get("description_enriched", "")
                                    or row.get("description", "") or ""),
                "footprint":    _fmt_footprint(fp_raw),
                # Display MPN: use real MPN; synthetic keys show blank
                "mpn":          "" if _is_synthetic(key) else key,
                "manufacturer": str(row.get("manufacturer", "") or ""),
                "mouser_pn":    str(row.get("mouser_pn",    "") or ""),
                "product_url":  str(row.get("product_url",  "") or ""),
                "unit_price":   _safe_float(row.get("unit_price")),
                "stock":        int(row.get("stock", 0) or 0),
                # For colour coding — use the first reference in the group
                "_ref_first":   ref,
                "_value":       str(row.get("value", "") or ""),
            }
            order.append(key)

    # ── Data rows ─────────────────────────────────────────────────────────────
    first_data_row = 4
    ext_col_letter = _EXT_COL   # K — used in TOTAL formula

    for offset, key in enumerate(order):
        excel_row = first_data_row + offset
        e         = seen[key]
        qty       = e["qty"]
        price     = e["unit_price"]

        u = f"{_UNIT_COL}{excel_row}"
        q = f"{_QTY_COL}{excel_row}"

        if price is not None:
            unit_gst = f"={u}*{gst_mult}"
            ext_base = f"={u}*{q}"
            ext_gst  = f"={u}*{q}*{gst_mult}"
        else:
            unit_gst = ext_base = ext_gst = None

        lead = _lead_time(e["stock"], qty)
        link = _purchase_url(e["product_url"], e["mouser_pn"])

        row_vals = [
            offset + 1,
            e["refs"],
            qty,
            e["description"],
            e["footprint"],
            e["mpn"],
            e["manufacturer"],
            e["mouser_pn"],
            price,
            unit_gst,
            ext_base,
            ext_gst,
            lead,
            link,
        ]

        # Row colour based on component type
        cat   = _category(e["_ref_first"], e["description"], e["_value"])
        dark, light = _CAT_COLORS[cat]
        bg    = dark if offset % 2 == 0 else light

        for ci, ((_, _, al), val) in enumerate(zip(_COLS, row_vals), 1):
            c           = ws.cell(row=excel_row, column=ci, value=val)
            c.font      = _font(size=9)
            c.alignment = _align(al, "center", wrap=(ci == 4))
            c.border    = _border()
            c.fill      = _fill(bg)

            if ci in (9, 10, 11, 12):
                c.number_format = '"₹"#,##0.00'
            elif ci == 14 and link:
                c.font = _font(size=9, color=_LINK_CLR)
                try:
                    ws.cell(row=excel_row, column=ci).hyperlink = link
                except Exception:
                    pass

        ws.row_dimensions[excel_row].height = 15

    last_data_row = first_data_row + len(order) - 1

    # ── TOTAL row ─────────────────────────────────────────────────────────────
    total_row = last_data_row + 2
    ws.merge_cells(f"A{total_row}:H{total_row}")
    t = ws[f"A{total_row}"]
    t.value     = "TOTAL COMPONENT COST (single unit)"
    t.font      = _font(size=9, bold=True)
    t.fill      = _fill(_NOTE_BG)
    t.alignment = _align("right", "center")

    # Ext. Price (₹) sum — col K
    ext_sum_col   = get_column_letter(11)   # K
    ext_gst_col   = get_column_letter(12)   # L

    for col_idx, col_letter in ((11, ext_sum_col), (12, ext_gst_col)):
        c = ws.cell(row=total_row, column=col_idx,
                    value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})")
        c.font          = _font(size=9, bold=True)
        c.fill          = _fill(_NOTE_BG)
        c.alignment     = _align("right", "center")
        c.number_format = '"₹"#,##0.00'
        c.border        = _border()

    # Fill I, J cols in total row (blank but styled)
    for col_idx in (9, 10):
        c = ws.cell(row=total_row, column=col_idx)
        c.fill   = _fill(_NOTE_BG)
        c.border = _border()

    ws.row_dimensions[total_row].height = 16

    # ── Colour legend ─────────────────────────────────────────────────────────
    legend_start = total_row + 2
    ws.merge_cells(f"A{legend_start}:{last_col}{legend_start}")
    h           = ws[f"A{legend_start}"]
    h.value     = "COLOUR LEGEND:"
    h.font      = _font(size=9, bold=True)
    h.alignment = _align("left", "center")
    ws.row_dimensions[legend_start].height = 14

    for i, (cat, text) in enumerate(_LEGEND, 1):
        r = legend_start + i
        ws.merge_cells(f"A{r}:{last_col}{r}")
        dark, _ = _CAT_COLORS[cat]
        c           = ws[f"A{r}"]
        c.value     = f"  {text}"
        c.font      = _font(size=9)
        c.fill      = _fill(dark)
        c.alignment = _align("left", "center")
        ws.row_dimensions[r].height = 14

    ws.freeze_panes = "A4"


# ── Public entry point ────────────────────────────────────────────────────────

def export_excel(df: pd.DataFrame, output_path: str,
                 project_name: str = "PCB Project") -> None:
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb       = Workbook()
    ws       = wb.active
    ws.title = "BOM"
    _write_sheet(ws, df, project_name)
    wb.save(output_path)
    print(f"  Saved → {output_path}")
